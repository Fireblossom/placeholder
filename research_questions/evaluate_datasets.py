#!/usr/bin/env python3
"""
Orthogonal and fair evaluation of extracted dataset lists.

Motivation
----------
Prior evaluations often conflate different constructs: coverage (aka recall-like
hit rate), redundancy, metadata completeness, and provenance/evidence. This
script implements a simple but principled protocol where these constructs are
orthogonal and complementary, inspired by FAIR (Wilkinson et al., Sci Data 2016),
ISO/IEC 25012 (data quality), W3C PROV-DM, and entity resolution practices
(Elmagarmid et al., 2007; Christen, 2012).

Key Principles
--------------
- Entity vs mention: All main metrics are computed on deduplicated entities
  (entity layer). Mention-level signals are reported separately for diagnostics.
- Evidence strength layering: We report coverage at three matcher strengths:
  Exact, Norm (mild normalization), and Fuzzy@tau; these are parallel views,
  not additive. For each, we also provide Evidence-backed and Trusted-backed
  slices (non-additive, to indicate confidence).
- Aggregation: Report per-file metrics, plus an aggregate recomputed from
  summed numerators/denominators (micro average). Macro averages can be derived
  externally if needed.

Main Metrics (entity layer unless noted)
---------------------------------------
- Coverage_T = |U_T ∩ G| / |G| for T ∈ {Exact, Norm, Fuzzy@tau}
- EBC_T (Evidence-backed Coverage) = |{e ∈ U_T ∩ G : has_any_evidence(e)}| / |G|
- TBC_T (Trusted-backed Coverage) = |{e ∈ U_T ∩ G : has_trusted_evidence(e)}| / |G|
- Redundancy (diagnostic, mention→entity) = (|mentions| − |U_Norm|) / max(1, |U_Norm|)
- Metadata completeness Meta@k (default k=3 over {title/author/year})
  on U_Norm
- PID Rate = |{e ∈ U_Norm : has_PID(e)}| / |U_Norm|
- Novelty_Norm = |U_Norm \ Baseline| / |U_Norm| if a baseline is provided; otherwise N/A

Inputs
------
- Directory containing one or more TSV/CSV files with header. Candidate columns
  for dataset names can be configured; we use the first present column.
- Optional gold file (txt/tsv/csv). If not provided, we look for gold.csv in
  the input directory. Gold is a set of names.
- Optional baseline file(s) for novelty.

Outputs
-------
- Per-file TSV with metrics listed below
- Aggregate TSV recomputed from numerators/denominators (micro average)

CLI Examples
------------
python evaluate_datasets_v2.py /path/to/dir --output-dir /path/to/out \
  --name-columns "Name;Dataset;Dataset Name" --citing-columns "Citing Article" \
  --cited-columns "Cited Article" --fuzzy-threshold 0.9

Notes
-----
- This script does NOT perform live URL availability checks by default, to avoid
  time-varying noise. Use --live-availability-sample to opt-in sampled checks.
- All evidence-based slices are orthogonal views and must not be added together.
"""

import argparse
import csv
import difflib
import os
import re
import sys
import urllib.parse
from collections import Counter, defaultdict
from typing import Dict, Iterable, List, Optional, Sequence, Set, Tuple

try:
    import requests  # type: ignore
except Exception:  # pragma: no cover
    requests = None

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover
    openpyxl = None


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Orthogonal evaluation of dataset extraction results",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument(
        "input_dir",
        nargs="?",
        default=os.path.dirname(__file__),
        help="Directory containing TSV/CSV/XLSX files or a single file (.tsv/.csv/.xlsx)",
    )
    parser.add_argument(
        "--pattern",
        default="*.tsv",
        help="Glob pattern for input files (TSV/CSV). XLSX files are always included.",
    )
    parser.add_argument(
        "--output-dir",
        default=os.path.join(os.path.dirname(__file__), "../evaluation_v2"),
        help="Directory to write reports",
    )
    parser.add_argument(
        "--name-columns",
        type=str,
        default="Name;Name (extracted);Dataset;Dataset Name",
        help="Semicolon-separated candidate columns for dataset names",
    )
    parser.add_argument(
        "--citing-columns",
        type=str,
        default="Citing Article;Citing_URL;Citing;Used in Which Papers",
        help="Semicolon-separated candidate columns for citing/source links",
    )
    parser.add_argument(
        "--cited-columns",
        type=str,
        default="Citied Article;Cited Article;Cited_URL;Cited;Introduced by Which Papers",
        help="Semicolon-separated candidate columns for cited links",
    )
    parser.add_argument(
        "--url-column",
        type=str,
        default="Dataset URL",
        help="Main dataset URL column (if present)",
    )
    # Metadata completeness disabled: field-map/metadata-k removed
    parser.add_argument(
        "--gold-file",
        type=str,
        default="",
        help=(
            "Gold file (txt/tsv/csv). If not set, uses gold.csv in input_dir if present; "
            "otherwise falls back to all inputs named 'survey' (sheets/files)."
        ),
    )
    parser.add_argument(
        "--gold-column",
        type=str,
        default="",
        help="Column name in gold file (if TSV/CSV)",
    )
    parser.add_argument(
        "--baseline",
        type=str,
        nargs="*",
        default=[],
        help="Baseline file(s) for novelty (txt/tsv/csv)",
    )
    parser.add_argument(
        "--baseline-column",
        type=str,
        default="",
        help="Column name in baseline files (if TSV/CSV)",
    )
    parser.add_argument(
        "--fuzzy-threshold",
        type=float,
        default=0.9,
        help="Similarity threshold for fuzzy matching (0-1)",
    )
    parser.add_argument(
        "--live-availability-sample",
        type=float,
        default=0.0,
        help="Sample ratio (0-1) for live URL availability checks (optional)",
    )
    parser.add_argument(
        "--check-urls",
        action="store_true",
        help="Enable live checking of dataset URLs for validity",
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=6.0,
        help="Per-request timeout seconds for live checks",
    )
    parser.add_argument(
        "--trust-hosts",
        type=str,
        default=(
            "doi.org;dx.doi.org;handle.net;hdl.handle.net;" \
            "pubmed.ncbi.nlm.nih.gov;ncbi.nlm.nih.gov;geo.ncbi.nlm.nih.gov;ebi.ac.uk;zenodo.org;figshare.com;" \
            "dataverse.org;datadryad.org;data.gov;kaggle.com;osf.io;openalex.org;openaire.eu;github.com;gitlab.com;" \
            "huggingface.co;opendatalab.com;scidb.cn;www.scidb.cn;ieee-dataport.org"
        ),
        help="Semicolon-separated trusted hostnames for provenance",
    )
    return parser.parse_args(argv)


def list_input_files(input_dir: str, pattern: str) -> List[str]:
    import glob

    search = os.path.join(input_dir, pattern)
    files = sorted(glob.glob(search))
    return [f for f in files if os.path.isfile(f)]


def read_header_sample(path: str) -> Tuple[List[str], str]:
    try:
        with open(path, "r", encoding="utf-8") as fh:
            sample = fh.read(4096)
    except Exception:
        return [], ""
    is_tsv = "\t" in sample
    is_csv = "," in sample and not is_tsv
    delim = "\t" if is_tsv else ("," if is_csv else "")
    if not delim:
        return ["_single"], ""
    reader = csv.reader(sample.splitlines(), delimiter=delim)
    try:
        header = next(reader)
    except StopIteration:
        header = []
    return [h.strip() for h in header], delim


def read_table(path: str) -> Tuple[List[Dict[str, str]], List[str]]:
    # Excel support: path can be "file.xlsx" or "file.xlsx#SheetName"
    if path.lower().endswith(".xlsx") or ".xlsx#" in path.lower():
        if openpyxl is None:
            raise RuntimeError("openpyxl is required to read .xlsx files. Please install it (e.g., pip install openpyxl).")

        def split_excel_path(p: str) -> Tuple[str, str]:
            # Split on the last '#', if present
            if "#" in p:
                base, sheet = p.rsplit("#", 1)
                return base, sheet
            return p, ""

        xlsx_path, sheet_name = split_excel_path(path)
        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        except FileNotFoundError:
            return [], []
        except Exception as exc:  # pragma: no cover
            raise RuntimeError(f"Failed to open Excel workbook: {xlsx_path}: {exc}")

        if sheet_name:
            if sheet_name not in wb.sheetnames:
                # Sheet missing -> empty
                return [], []
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # Extract header from the first non-empty row
        headers: List[str] = []
        rows_iter = ws.iter_rows(values_only=True)
        first_row: Optional[Tuple[object, ...]] = None
        for r in rows_iter:
            if any(c is not None and str(c).strip() != "" for c in r):
                first_row = r  # type: ignore[assignment]
                break
        if first_row is None:
            return [], []

        # Build headers, ensuring non-empty unique names
        seen: Set[str] = set()
        tmp_headers: List[str] = []
        for i, val in enumerate(first_row):
            name = str(val).strip() if val is not None else ""
            if not name:
                name = f"Column{i+1}"
            base = name
            j = 1
            while name in seen:
                j += 1
                name = f"{base}_{j}"
            seen.add(name)
            tmp_headers.append(name)
        headers = tmp_headers

        # Remaining rows become data
        data_rows: List[Dict[str, str]] = []
        for r in rows_iter:
            # Normalize row length to headers
            vals = list(r)
            if len(vals) < len(headers):
                vals.extend([None] * (len(headers) - len(vals)))
            elif len(vals) > len(headers):
                vals = vals[: len(headers)]
            # Skip completely empty rows
            if not any(c is not None and str(c).strip() != "" for c in vals):
                continue
            rec: Dict[str, str] = {}
            for h, v in zip(headers, vals):
                if v is None:
                    rec[h] = ""
                else:
                    s = str(v)
                    rec[h] = s.strip()
            data_rows.append(rec)
        return data_rows, headers

    headers, delim = read_header_sample(path)
    if not delim:
        # treat as single-column text
        rows: List[Dict[str, str]] = []
        with open(path, "r", encoding="utf-8") as fh:
            for line in fh:
                rows.append({headers[0]: line.strip()})
        return rows, [headers[0]]
    with open(path, "r", encoding="utf-8") as fh:
        reader = csv.DictReader(fh, delimiter=delim)
        rows = [dict(r) for r in reader]
        return rows, reader.fieldnames or []


def normalize_header_key(s: str) -> str:
    s2 = s.strip().lower()
    s2 = s2.replace("_", " ")
    s2 = re.sub(r"\s+", " ", s2)
    return s2


def pick_first_column(candidates: Sequence[str], headers: Sequence[str]) -> Optional[str]:
    if not headers:
        return None
    norm_map = {normalize_header_key(h): h for h in headers}
    for cand in candidates:
        key = normalize_header_key(cand)
        if key in norm_map:
            return norm_map[key]
    return None


def extract_names(rows: List[Dict[str, str]], name_columns: Sequence[str]) -> List[str]:
    if not rows:
        return []
    headers = list(rows[0].keys())
    picked = pick_first_column(name_columns, headers)
    if not picked:
        return []
    names: List[str] = []
    for r in rows:
        names.append((r.get(picked, "") or "").strip())
    return names


def canonical_exact(name: str) -> str:
    if not name:
        return ""
    s = re.sub(r"\s+", " ", name.strip())
    return s


def canonical_norm(name: str) -> str:
    if not name:
        return ""
    s = name.strip().lower()
    s = re.sub(r"[\s\u00A0]+", " ", s)
    # remove light punctuation only; keep parentheses/hyphens content
    s = re.sub(r"[\"'`“”‘’.,;:]+", "", s)
    s = re.sub(r"\s+", " ", s)
    return s


def fuzzy_key(name: str) -> str:
    # prepare a simplified key for fuzzy similarity
    s = canonical_norm(name)
    s = re.sub(r"[^a-z0-9 ]", "", s)
    return s


def cluster_fuzzy(names: List[str], threshold: float) -> List[List[str]]:
    """Greedy single-link clustering over names using difflib ratio on fuzzy_key.
    Returns list of clusters where each cluster is a list of original names.
    """
    if not names:
        return []
    uniq = []
    seen: Set[str] = set()
    for n in names:
        if n not in seen:
            seen.add(n)
            uniq.append(n)
    reps: List[str] = []
    clusters: List[List[str]] = []
    for n in uniq:
        k = fuzzy_key(n)
        assigned = False
        for i, rep in enumerate(reps):
            kr = fuzzy_key(rep)
            sim = difflib.SequenceMatcher(None, k, kr).ratio()
            if sim >= threshold:
                clusters[i].append(n)
                assigned = True
                break
        if not assigned:
            reps.append(n)
            clusters.append([n])
    return clusters


def parse_hosts(trust_hosts: str) -> Set[str]:
    hosts = set()
    for item in trust_hosts.split(";"):
        host = item.strip().lower()
        if host:
            hosts.add(host)
    return hosts


DOI_RE = re.compile(r"\b10\.\d{4,9}/\S+\b", re.IGNORECASE)
HANDLE_RE = re.compile(r"\b(?:hdl\.)?handle\.net/\S+\b", re.IGNORECASE)
ARK_RE = re.compile(r"\bark:/\S+\b", re.IGNORECASE)


def extract_urls_from_row(row: Dict[str, str], columns: Sequence[str]) -> List[str]:
    urls: List[str] = []
    for c in columns:
        v = (row.get(c, "") or "").strip()
        if v:
            urls.append(v)
    return urls


def url_host(url: str) -> str:
    try:
        if not re.match(r"^[a-zA-Z][a-zA-Z0-9+.-]*://", url):
            url = "https://" + url
        parsed = urllib.parse.urlparse(url)
        host = parsed.netloc.lower()
        return host
    except Exception:
        return ""


def has_any_evidence(urls: Iterable[str]) -> bool:
    for u in urls:
        if (u or "").strip():
            return True
    return False


def has_trusted_evidence(urls: Iterable[str], trusted_hosts: Set[str]) -> bool:
    for u in urls:
        if not u:
            continue
        if DOI_RE.search(u) or HANDLE_RE.search(u) or ARK_RE.search(u):
            return True
        h = url_host(u)
        if h and (h in trusted_hosts or any(h.endswith("." + th) for th in trusted_hosts)):
            return True
    return False


def has_pid(urls: Iterable[str]) -> bool:
    for u in urls:
        if not u:
            continue
        if DOI_RE.search(u) or HANDLE_RE.search(u) or ARK_RE.search(u):
            return True
    return False


def is_url_working(url: str, timeout: float) -> bool:
    if not requests:
        return False
    if not url:
        return False
    try:
        u = url
        if not re.match(r"^[a-zA-Z][a-zA-Z0-9+.-]*://", u):
            u = "https://" + u
        resp = requests.head(u, allow_redirects=True, timeout=timeout)
        if 200 <= resp.status_code < 300:
            return True
        resp = requests.get(u, allow_redirects=True, timeout=timeout, stream=True)
        try:
            return 200 <= resp.status_code < 300
        finally:
            try:
                resp.close()
            except Exception:
                pass
    except Exception:
        return False


def autodetect_field_map(headers: Sequence[str]) -> Dict[str, str]:
    # Metadata completeness disabled; keep placeholder for backward compatibility
    return {}


def compute_meta_at_k(entity_rows: List[Dict[str, str]], field_map: Dict[str, str], k: int) -> Tuple[int, int]:
    # Metadata completeness disabled
    return 0, 0


def load_name_list_from_file(path: str, column: str = "") -> List[str]:
    names: List[str] = []
    if not path:
        return names
    try:
        with open(path, "r", encoding="utf-8") as fh:
            sample = fh.read(4096)
            fh.seek(0)
            is_tsv = "\t" in sample
            is_csv = "," in sample and not is_tsv
            if is_tsv or is_csv:
                delim = "\t" if is_tsv else ","
                reader = csv.DictReader(fh, delimiter=delim)
                if column and column in (reader.fieldnames or []):
                    for r in reader:
                        nm = (r.get(column, "") or "").strip()
                        if nm:
                            names.append(nm)
                else:
                    first = (reader.fieldnames or [""])[0]
                    for r in reader:
                        nm = (r.get(first, "") or "").strip()
                        if nm:
                            names.append(nm)
            else:
                for line in fh:
                    nm = line.strip()
                    if nm:
                        names.append(nm)
    except FileNotFoundError:
        return []
    return names


def build_entities(
    rows: List[Dict[str, str]],
    name_columns: Sequence[str],
    url_column: str,
    citing_columns: Sequence[str],
    cited_columns: Sequence[str],
    matcher: str,
    fuzzy_threshold: float,
) -> Tuple[List[Dict[str, object]], List[str]]:
    """Return (entities, mentions) where entities are deduplicated group dicts.
    entity dict includes: names, rows_idx, evidence_urls, repr_name.
    matcher in {Exact, Norm, Fuzzy}.
    """
    names = extract_names(rows, name_columns)
    mentions: List[str] = [n for n in names if (n or "").strip()]
    if not mentions:
        return [], []

    # Gather per-mention evidence and metadata view (row index -> urls)
    headers = list(rows[0].keys()) if rows else []
    url_cols: List[str] = []
    # prioritize declared url_column if exists
    if url_column and url_column in headers:
        url_cols.append(url_column)
    for c in citing_columns:
        if c in headers:
            url_cols.append(c)
    for c in cited_columns:
        if c in headers:
            url_cols.append(c)

    def entity_from_indices(idxs: List[int]) -> Dict[str, object]:
        ev: List[str] = []
        home: List[str] = []
        for i in idxs:
            ev.extend(extract_urls_from_row(rows[i], url_cols))
            if url_column and url_column in rows[i]:
                u = (rows[i].get(url_column, "") or "").strip()
                if u:
                    home.append(u)
        rep = mentions[idxs[0]]
        return {
            "names": [mentions[i] for i in idxs],
            "rows_idx": idxs,
            "evidence_urls": [u for u in ev if (u or "").strip()],
            "dataset_urls": [u for u in home if (u or "").strip()],
            "repr_name": rep,
        }

    if matcher == "Exact":
        key_to_indices: Dict[str, List[int]] = defaultdict(list)
        for i, n in enumerate(mentions):
            key = canonical_exact(n)
            key_to_indices[key].append(i)
        entities = [entity_from_indices(v) for v in key_to_indices.values()]
        return entities, mentions
    elif matcher == "Norm":
        key_to_indices = defaultdict(list)
        for i, n in enumerate(mentions):
            key = canonical_norm(n)
            key_to_indices[key].append(i)
        entities = [entity_from_indices(v) for v in key_to_indices.values()]
        return entities, mentions
    else:
        # Fuzzy clustering on fuzzy_key
        clusters = cluster_fuzzy(mentions, fuzzy_threshold)
        entities: List[Dict[str, object]] = []
        name_to_indices: Dict[str, List[int]] = defaultdict(list)
        for idx, n in enumerate(mentions):
            name_to_indices[n].append(idx)
        for cl in clusters:
            idxs: List[int] = []
            for n in cl:
                idxs.extend(name_to_indices.get(n, []))
            idxs.sort()
            entities.append(entity_from_indices(idxs))
        return entities, mentions


def compute_coverage(
    entities: List[Dict[str, object]],
    gold_names: List[str],
    matcher: str,
    fuzzy_threshold: float,
) -> Tuple[int, int, int, int]:
    """Return tuple: (hit_entities, gold_total, ebc_hit_entities, tbc_hit_entities).
    Evidence-based hits are restricted to entities that are hits vs gold.
    """
    gold_total = len({canonical_exact(g) for g in gold_names}) if matcher == "Exact" else (
        len({canonical_norm(g) for g in gold_names}) if matcher == "Norm" else len(cluster_fuzzy(gold_names, fuzzy_threshold))
    )
    if gold_total == 0:
        return 0, 0, 0, 0

    # Prepare gold views
    if matcher == "Exact":
        gold_keys = {canonical_exact(g) for g in gold_names}
        def entity_key(e: Dict[str, object]) -> str:
            return canonical_exact(str(e.get("repr_name", "")))
    elif matcher == "Norm":
        gold_keys = {canonical_norm(g) for g in gold_names}
        def entity_key(e: Dict[str, object]) -> str:
            return canonical_norm(str(e.get("repr_name", "")))
    else:
        # For fuzzy, use clustered survey keys as gold representatives
        clusters = cluster_fuzzy(gold_names, fuzzy_threshold)
        gold_list = [fuzzy_key(cl[0]) for cl in clusters if cl]

        def entity_key(e: Dict[str, object]) -> str:
            return fuzzy_key(str(e.get("repr_name", "")))

    trusted_hosts: Set[str] = set()  # placeholder, will be supplied externally if needed
    # We leave evidence slices to caller variants that have trusted_hosts
    raise_on_trusted = False
    if raise_on_trusted:  # pragma: no cover
        pass

    hit = 0
    ebc_hit = 0
    tbc_hit = 0
    # For evidence slices we will compute externally; here we only compute hit counts
    if matcher in {"Exact", "Norm"}:
        gold_keys_set = set(gold_keys)  # type: ignore
        for e in entities:
            k = entity_key(e)
            if k in gold_keys_set:
                hit += 1
    else:
        for e in entities:
            k = entity_key(e)
            best = 0.0
            for gg in gold_list:
                sim = difflib.SequenceMatcher(None, k, gg).ratio()
                if sim > best:
                    best = sim
                    if best >= fuzzy_threshold:
                        break
            if best >= fuzzy_threshold:
                hit += 1
    return hit, gold_total, ebc_hit, tbc_hit


def compute_evidence_slices(
    entities: List[Dict[str, object]],
    gold_names: List[str],
    matcher: str,
    fuzzy_threshold: float,
    trusted_hosts: Set[str],
) -> Tuple[int, int, int, int, int]:
    """Return tuple restricted to entities that are hits vs gold:
    (EBC_hit, TBC_hit, Hit_with_DatasetURL, Hit_with_Working_DatasetURL, Hit_total)
    Note: Working dataset URL requires --check-urls; otherwise it remains 0.
    """
    if not gold_names:
        return 0, 0, 0, 0, 0

    # Prepare gold structures
    if matcher == "Exact":
        gold_keys = {canonical_exact(g) for g in gold_names}
        def is_hit(name: str) -> bool:
            return canonical_exact(name) in gold_keys
    elif matcher == "Norm":
        gold_keys = {canonical_norm(g) for g in gold_names}
        def is_hit(name: str) -> bool:
            return canonical_norm(name) in gold_keys
    else:
        gold_list = [fuzzy_key(g) for g in gold_names]
        def is_hit(name: str) -> bool:
            k = fuzzy_key(name)
            best = 0.0
            for gg in gold_list:
                sim = difflib.SequenceMatcher(None, k, gg).ratio()
                if sim > best:
                    best = sim
                    if best >= fuzzy_threshold:
                        break
            return best >= fuzzy_threshold

    ebc = 0
    tbc = 0
    hit_total = 0
    hit_with_dataset_url = 0
    hit_with_working_dataset_url = 0
    for e in entities:
        rep = e.get("repr_name", "") or ""
        if not isinstance(rep, str) or not is_hit(rep):
            continue
        hit_total += 1
        urls = e.get("evidence_urls", [])
        urls_list = urls if isinstance(urls, list) else []
        if has_any_evidence(urls_list):
            ebc += 1
        if has_trusted_evidence(urls_list, trusted_hosts):
            tbc += 1
        homes = e.get("dataset_urls", [])
        homes_list = homes if isinstance(homes, list) else []
        if any((u or "").strip() for u in homes_list):
            hit_with_dataset_url += 1
        # working dataset URL will be computed in caller depending on --check-urls
    return ebc, tbc, hit_with_dataset_url, hit_with_working_dataset_url, hit_total


def compute_redundancy(num_mentions: int, num_entities_norm: int) -> float:
    if num_entities_norm <= 0:
        return 0.0
    return max(0.0, float(num_mentions - num_entities_norm) / float(num_entities_norm))


def percent(numer: int, denom: int) -> float:
    if denom <= 0:
        return 0.0
    return 100.0 * float(numer) / float(denom)


def evaluate_file(
    path: str,
    args: argparse.Namespace,
    gold_names: List[str],
    baseline_names: Set[str],
    trusted_hosts: Set[str],
) -> Dict[str, object]:
    rows, headers = read_table(path)
    name_columns = [c.strip() for c in args.name_columns.split(";") if c.strip()]
    citing_cols = [c.strip() for c in args.citing_columns.split(";") if c.strip()]
    cited_cols = [c.strip() for c in args.cited_columns.split(";") if c.strip()]

    # Entities under three matchers
    ent_exact, mentions = build_entities(
        rows, name_columns, args.url_column, citing_cols, cited_cols, "Exact", args.fuzzy_threshold
    )
    ent_norm, _ = build_entities(
        rows, name_columns, args.url_column, citing_cols, cited_cols, "Norm", args.fuzzy_threshold
    )
    ent_fuzzy, _ = build_entities(
        rows, name_columns, args.url_column, citing_cols, cited_cols, "Fuzzy", args.fuzzy_threshold
    )

    mentions_count = len(mentions)
    entities_exact = len(ent_exact)
    entities_norm = len(ent_norm)
    entities_fuzzy = len(ent_fuzzy)

    # Coverage and evidence slices
    cov_e_hit, cov_e_total, _, _ = compute_coverage(ent_exact, gold_names, "Exact", args.fuzzy_threshold)
    ebc_e, tbc_e, hit_home_e, hit_working_home_e, hit_total_e = compute_evidence_slices(
        ent_exact, gold_names, "Exact", args.fuzzy_threshold, trusted_hosts
    )

    cov_n_hit, cov_n_total, _, _ = compute_coverage(ent_norm, gold_names, "Norm", args.fuzzy_threshold)
    ebc_n, tbc_n, hit_home_n, hit_working_home_n, hit_total_n = compute_evidence_slices(
        ent_norm, gold_names, "Norm", args.fuzzy_threshold, trusted_hosts
    )

    cov_f_hit, cov_f_total, _, _ = compute_coverage(ent_fuzzy, gold_names, "Fuzzy", args.fuzzy_threshold)
    ebc_f, tbc_f, hit_home_f, hit_working_home_f, hit_total_f = compute_evidence_slices(
        ent_fuzzy, gold_names, "Fuzzy", args.fuzzy_threshold, trusted_hosts
    )

    # Optionally compute working dataset URL counts via live checks
    def compute_working_home_hits(entities: List[Dict[str, object]], matcher: str) -> int:
        if not args.check_urls or not gold_names:
            return 0
        # define is_hit consistent with matcher
        if matcher == "Exact":
            gold_keys = {canonical_exact(g) for g in gold_names}
            def is_hit_name(n: str) -> bool:
                return canonical_exact(n) in gold_keys
        elif matcher == "Norm":
            gold_keys = {canonical_norm(g) for g in gold_names}
            def is_hit_name(n: str) -> bool:
                return canonical_norm(n) in gold_keys
        else:
            gold_list = [fuzzy_key(g) for g in gold_names]
            def is_hit_name(n: str) -> bool:
                k = fuzzy_key(n)
                best = 0.0
                for gg in gold_list:
                    sim = difflib.SequenceMatcher(None, k, gg).ratio()
                    if sim > best:
                        best = sim
                        if best >= args.fuzzy_threshold:
                            break
                return best >= args.fuzzy_threshold

        working_cnt = 0
        for e in entities:
            rep = e.get("repr_name", "") or ""
            if not isinstance(rep, str) or not is_hit_name(rep):
                continue
            homes = e.get("dataset_urls", [])
            homes_list = homes if isinstance(homes, list) else []
            ok = False
            for u in homes_list:
                if is_url_working(u, args.timeout):
                    ok = True
                    break
            if ok:
                working_cnt += 1
        return working_cnt

    if args.check_urls:
        hit_working_home_e = compute_working_home_hits(ent_exact, "Exact")
        hit_working_home_n = compute_working_home_hits(ent_norm, "Norm")
        hit_working_home_f = compute_working_home_hits(ent_fuzzy, "Fuzzy")

    # Redundancy (mention→entity under Norm)
    redundancy = compute_redundancy(mentions_count, entities_norm)

    # Metadata completeness disabled: meta_ok/meta_total set to 0
    meta_ok, meta_total = 0, 0

    # Evidence distributions (exclusive precedence: DOI/PID > TrustedHost > OtherLink > None)
    cat_counts = Counter()
    for e in ent_norm:
        urls = e.get("evidence_urls", [])
        urls_list = urls if isinstance(urls, list) else []
        if has_pid(urls_list):
            cat_counts["PID"] += 1
        elif has_trusted_evidence(urls_list, trusted_hosts):
            cat_counts["TrustedHost"] += 1
        elif has_any_evidence(urls_list):
            cat_counts["OtherLink"] += 1
        else:
            cat_counts["None"] += 1

    pid_rate_numer = cat_counts["PID"]
    pid_rate_denom = max(1, entities_norm)

    # Novelty on U_Norm
    novelty_numer = 0
    novelty_denom = entities_norm
    if baseline_names:
        baseline_set_norm = {canonical_norm(n) for n in baseline_names}
        for e in ent_norm:
            rep = e.get("repr_name", "") or ""
            repn = canonical_norm(rep if isinstance(rep, str) else str(rep))
            if repn not in baseline_set_norm:
                novelty_numer += 1

    # Derive method from file name suffix after '#', if present
    _file_base = os.path.basename(path)
    _method = ""
    if "#" in _file_base:
        _method = _file_base.rsplit("#", 1)[-1].strip()
    _display_file = _file_base.split("#", 1)[0]

    summary: Dict[str, object] = {
        "File": _display_file,
        "Method": _method,
        "Mentions": mentions_count,
        "Entities_Exact": entities_exact,
        "Entities_Norm": entities_norm,
        "Entities_Fuzzy": entities_fuzzy,
        # Coverage
        "Coverage_Exact_Hit": cov_e_hit,
        "Coverage_Exact_Total": cov_e_total,
        "Coverage_Exact_percent": round(percent(cov_e_hit, cov_e_total), 2),
        # Recall aliases and slices (Exact)
        "Recall_Exact_percent": round(percent(cov_e_hit, cov_e_total), 2),
        "EvidenceBacked_Recall_Exact_percent": round(percent(ebc_e, cov_e_total), 2),
        "TrustedBacked_Recall_Exact_percent": round(percent(tbc_e, cov_e_total), 2),
        "Recall_withDatasetURL_Exact_percent": round(percent(hit_home_e, cov_e_total), 2),
        "Recall_withValidDatasetURL_Exact_percent": round(percent(hit_working_home_e, cov_e_total), 2),
        "Hit_Exact_Total": hit_total_e,
        "Hit_Exact_WithDatasetURL": hit_home_e,
        "Hit_Exact_WithWorkingDatasetURL": hit_working_home_e,
        "EBC_Exact_Hit": ebc_e,
        "EBC_Exact_percent": round(percent(ebc_e, cov_e_total), 2),
        "TBC_Exact_Hit": tbc_e,
        "TBC_Exact_percent": round(percent(tbc_e, cov_e_total), 2),

        "Coverage_Norm_Hit": cov_n_hit,
        "Coverage_Norm_Total": cov_n_total,
        "Coverage_Norm_percent": round(percent(cov_n_hit, cov_n_total), 2),
        # Recall aliases and slices (Norm)
        "Recall_Norm_percent": round(percent(cov_n_hit, cov_n_total), 2),
        "EvidenceBacked_Recall_Norm_percent": round(percent(ebc_n, cov_n_total), 2),
        "TrustedBacked_Recall_Norm_percent": round(percent(tbc_n, cov_n_total), 2),
        "Recall_withDatasetURL_Norm_percent": round(percent(hit_home_n, cov_n_total), 2),
        "Recall_withValidDatasetURL_Norm_percent": round(percent(hit_working_home_n, cov_n_total), 2),
        "Hit_Norm_Total": hit_total_n,
        "Hit_Norm_WithDatasetURL": hit_home_n,
        "Hit_Norm_WithWorkingDatasetURL": hit_working_home_n,
        "EBC_Norm_Hit": ebc_n,
        "EBC_Norm_percent": round(percent(ebc_n, cov_n_total), 2),
        "TBC_Norm_Hit": tbc_n,
        "TBC_Norm_percent": round(percent(tbc_n, cov_n_total), 2),

        "Coverage_Fuzzy_Hit": cov_f_hit,
        "Coverage_Fuzzy_Total": cov_f_total,
        "Coverage_Fuzzy_percent": round(percent(cov_f_hit, cov_f_total), 2),
        # Recall aliases and slices (Fuzzy)
        "Recall_Fuzzy_percent": round(percent(cov_f_hit, cov_f_total), 2),
        "EvidenceBacked_Recall_Fuzzy_percent": round(percent(ebc_f, cov_f_total), 2),
        "TrustedBacked_Recall_Fuzzy_percent": round(percent(tbc_f, cov_f_total), 2),
        "Recall_withDatasetURL_Fuzzy_percent": round(percent(hit_home_f, cov_f_total), 2),
        "Recall_withValidDatasetURL_Fuzzy_percent": round(percent(hit_working_home_f, cov_f_total), 2),
        "Hit_Fuzzy_Total": hit_total_f,
        "Hit_Fuzzy_WithDatasetURL": hit_home_f,
        "Hit_Fuzzy_WithWorkingDatasetURL": hit_working_home_f,
        "EBC_Fuzzy_Hit": ebc_f,
        "EBC_Fuzzy_percent": round(percent(ebc_f, cov_f_total), 2),
        "TBC_Fuzzy_Hit": tbc_f,
        "TBC_Fuzzy_percent": round(percent(tbc_f, cov_f_total), 2),

        # Diagnostics and quality
        "Redundancy_rate": round(redundancy, 4),
        # Metadata completeness removed

        # Evidence distribution (entity layer, Norm)
        "Evidence_PID": cat_counts["PID"],
        "Evidence_TrustedHost": cat_counts["TrustedHost"],
        "Evidence_OtherLink": cat_counts["OtherLink"],
        "Evidence_None": cat_counts["None"],
        "PID_Rate_percent": round(percent(pid_rate_numer, pid_rate_denom), 2),

        # Novelty
        "Novel_Norm": novelty_numer,
        "Novel_Norm_Base": novelty_denom,
        "Novelty_Norm_percent": (round(percent(novelty_numer, novelty_denom), 2) if baseline_names else "N/A"),
    }
    return summary


def write_tsv(path: str, rows: List[Dict[str, object]], field_order: Optional[List[str]] = None) -> None:
    if not rows:
        with open(path, "w", encoding="utf-8") as fh:
            fh.write("")
        return
    fields: List[str] = field_order or list(rows[0].keys())
    # If Method exists but not immediately after File, move it to second column
    if "File" in fields and "Method" in fields:
        if fields.index("Method") != fields.index("File") + 1:
            fields.remove("Method")
            insert_at = fields.index("File") + 1
            fields.insert(insert_at, "Method")
    with open(path, "w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields, delimiter="\t")
        writer.writeheader()
        for r in rows:
            writer.writerow({k: r.get(k, "") for k in fields})


def aggregate_micro(summaries: List[Dict[str, object]]) -> Dict[str, object]:
    agg: Dict[str, float] = defaultdict(float)
    for s in summaries:
        for k, v in s.items():
            if k == "File":
                continue
            if isinstance(v, (int, float)):
                agg[k] += float(v)

    def pct(nk: str, dk: str) -> float:
        return percent(int(agg.get(nk, 0)), int(agg.get(dk, 0)))

    return {
        "Scope": "AGGREGATE_V2",
        "Mentions": int(agg.get("Mentions", 0)),
        "Entities_Exact": int(agg.get("Entities_Exact", 0)),
        "Entities_Norm": int(agg.get("Entities_Norm", 0)),
        "Entities_Fuzzy": int(agg.get("Entities_Fuzzy", 0)),

        "Coverage_Exact_Hit": int(agg.get("Coverage_Exact_Hit", 0)),
        "Coverage_Exact_Total": int(agg.get("Coverage_Exact_Total", 0)),
        "Coverage_Exact_percent": round(pct("Coverage_Exact_Hit", "Coverage_Exact_Total"), 2),
        "Recall_Exact_percent": round(pct("Coverage_Exact_Hit", "Coverage_Exact_Total"), 2),
        "EvidenceBacked_Recall_Exact_percent": round(pct("EBC_Exact_Hit", "Coverage_Exact_Total"), 2),
        "TrustedBacked_Recall_Exact_percent": round(pct("TBC_Exact_Hit", "Coverage_Exact_Total"), 2),
        "EBC_Exact_Hit": int(agg.get("EBC_Exact_Hit", 0)),
        "EBC_Exact_percent": round(pct("EBC_Exact_Hit", "Coverage_Exact_Total"), 2),
        "TBC_Exact_Hit": int(agg.get("TBC_Exact_Hit", 0)),
        "TBC_Exact_percent": round(pct("TBC_Exact_Hit", "Coverage_Exact_Total"), 2),

        "Coverage_Norm_Hit": int(agg.get("Coverage_Norm_Hit", 0)),
        "Coverage_Norm_Total": int(agg.get("Coverage_Norm_Total", 0)),
        "Coverage_Norm_percent": round(pct("Coverage_Norm_Hit", "Coverage_Norm_Total"), 2),
        "Recall_Norm_percent": round(pct("Coverage_Norm_Hit", "Coverage_Norm_Total"), 2),
        "EvidenceBacked_Recall_Norm_percent": round(pct("EBC_Norm_Hit", "Coverage_Norm_Total"), 2),
        "TrustedBacked_Recall_Norm_percent": round(pct("TBC_Norm_Hit", "Coverage_Norm_Total"), 2),
        "EBC_Norm_Hit": int(agg.get("EBC_Norm_Hit", 0)),
        "EBC_Norm_percent": round(pct("EBC_Norm_Hit", "Coverage_Norm_Total"), 2),
        "TBC_Norm_Hit": int(agg.get("TBC_Norm_Hit", 0)),
        "TBC_Norm_percent": round(pct("TBC_Norm_Hit", "Coverage_Norm_Total"), 2),

        "Coverage_Fuzzy_Hit": int(agg.get("Coverage_Fuzzy_Hit", 0)),
        "Coverage_Fuzzy_Total": int(agg.get("Coverage_Fuzzy_Total", 0)),
        "Coverage_Fuzzy_percent": round(pct("Coverage_Fuzzy_Hit", "Coverage_Fuzzy_Total"), 2),
        "Recall_Fuzzy_percent": round(pct("Coverage_Fuzzy_Hit", "Coverage_Fuzzy_Total"), 2),
        "EvidenceBacked_Recall_Fuzzy_percent": round(pct("EBC_Fuzzy_Hit", "Coverage_Fuzzy_Total"), 2),
        "TrustedBacked_Recall_Fuzzy_percent": round(pct("TBC_Fuzzy_Hit", "Coverage_Fuzzy_Total"), 2),
        "EBC_Fuzzy_Hit": int(agg.get("EBC_Fuzzy_Hit", 0)),
        "EBC_Fuzzy_percent": round(pct("EBC_Fuzzy_Hit", "Coverage_Fuzzy_Total"), 2),
        "TBC_Fuzzy_Hit": int(agg.get("TBC_Fuzzy_Hit", 0)),
        "TBC_Fuzzy_percent": round(pct("TBC_Fuzzy_Hit", "Coverage_Fuzzy_Total"), 2),

        "Redundancy_rate": round(float(agg.get("Redundancy_rate", 0.0)) / max(1.0, float(len(summaries))), 4),

        "Evidence_PID": int(agg.get("Evidence_PID", 0)),
        "Evidence_TrustedHost": int(agg.get("Evidence_TrustedHost", 0)),
        "Evidence_OtherLink": int(agg.get("Evidence_OtherLink", 0)),
        "Evidence_None": int(agg.get("Evidence_None", 0)),
        "PID_Rate_percent": round(float(agg.get("PID_Rate_percent", 0.0)) / max(1.0, float(len(summaries))), 2),

        "Novel_Norm": int(agg.get("Novel_Norm", 0)),
        "Novel_Norm_Base": int(agg.get("Novel_Norm_Base", 0)),
        "Novelty_Norm_percent": (
            round(pct("Novel_Norm", "Novel_Norm_Base"), 2) if int(agg.get("Novel_Norm_Base", 0)) > 0 else "N/A"
        ),
    }


def detect_gold_csv(input_dir: str) -> Optional[str]:
    gold_path = os.path.join(input_dir, "gold.csv")
    if os.path.isfile(gold_path):
        return gold_path
    return None


def is_survey_input(path: str) -> bool:
    base = os.path.basename(path)
    lower = base.lower()
    return lower.endswith("#survey") or re.search(r"(^|[.#])survey(\.|$)", lower) is not None


def load_survey_as_gold(files: List[str], gold_column: str) -> List[str]:
    # Treat any input ending with '#survey' (xlsx sheet) or file named 'survey.tsv/csv' as gold
    names: List[str] = []
    for f in files:
        if is_survey_input(f):
            rows, _ = read_table(f)
            # Use current name-columns logic to extract names
            # We need to approximate since parse_args isn't available here; caller will pass names via standard flow
            # Instead, try common headers
            headers = rows[0].keys() if rows else []
            for cand in ["Name", "Name (extracted)", "Dataset", "Dataset Name"]:
                if cand in headers:
                    for r in rows:
                        nm = (r.get(cand, "") or "").strip()
                        if nm:
                            names.append(nm)
                    break
    return names


def get_workbook_base(path: str) -> str:
    # Return the base .xlsx path if this is an xlsx sheet input, else ""
    p = path
    if ".xlsx#" in p.lower() and p.lower().endswith(
        p.split("#", 1)[-1].lower()
    ):
        return p.rsplit("#", 1)[0]
    return ""


def build_per_workbook_gold(files: List[str], name_columns: Sequence[str]) -> Dict[str, List[str]]:
    per_gold: Dict[str, List[str]] = {}
    # Find survey sheet per workbook and extract names
    for f in files:
        base = get_workbook_base(f)
        if not base:
            continue
        if is_survey_input(f):
            rows, _ = read_table(f)
            if not rows:
                per_gold[base] = []
                continue
            names = extract_names(rows, name_columns)
            per_gold[base] = [n for n in names if (n or "").strip()]
    return per_gold


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    os.makedirs(args.output_dir, exist_ok=True)

    # Find input files or Excel sheets
    def expand_inputs(input_arg: str, pattern: str) -> List[str]:
        # If a direct file is provided
        if os.path.isfile(input_arg):
            if input_arg.lower().endswith(".xlsx"):
                if openpyxl is None:
                    raise RuntimeError("openpyxl is required to read .xlsx files. Please install it (e.g., pip install openpyxl).")
                try:
                    wb = openpyxl.load_workbook(input_arg, data_only=True, read_only=True)
                except Exception as exc:  # pragma: no cover
                    raise RuntimeError(f"Failed to open Excel workbook: {input_arg}: {exc}")
                return [f"{input_arg}#{sn}" for sn in wb.sheetnames]
            # Non-Excel single file
            return [input_arg]
        # Otherwise treat as a directory
        files = list_input_files(input_arg, pattern)
        # Always include XLSX files regardless of --pattern
        try:
            import glob as _glob  # local import to avoid top-level changes
            xlsx_files = set()
            for _pat in ("*.xlsx", "*.XLSX"):
                for _f in _glob.glob(os.path.join(input_arg, _pat)):
                    if os.path.isfile(_f):
                        xlsx_files.add(_f)
            for _xf in sorted(xlsx_files):
                if _xf not in files:
                    files.append(_xf)
        except Exception:
            pass
        expanded: List[str] = []
        for f in files:
            if f.lower().endswith(".xlsx"):
                if openpyxl is None:
                    raise RuntimeError("openpyxl is required to read .xlsx files. Please install it (e.g., pip install openpyxl).")
                try:
                    wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
                except Exception as exc:  # pragma: no cover
                    raise RuntimeError(f"Failed to open Excel workbook: {f}: {exc}")
                expanded.extend([f"{f}#{sn}" for sn in wb.sheetnames])
            else:
                expanded.append(f)
        return expanded

    files = expand_inputs(args.input_dir, args.pattern)
    if not files:
        print("No input files found.", file=sys.stderr)
        return 2

    # Gold
    gold_names: List[str] = []
    if args.gold_file:
        gold_names = load_name_list_from_file(args.gold_file, args.gold_column)
        print(f"Using explicit gold file: {os.path.basename(args.gold_file)}")
    else:
        gold_path = detect_gold_csv(args.input_dir)
        if gold_path:
            gold_names = load_name_list_from_file(gold_path, args.gold_column)
            print(f"Found gold.csv in input directory: {os.path.basename(gold_path)}")
        # If not using explicit or gold.csv, we'll do per-workbook survey-as-gold later

    # Baseline (novelty)
    baseline_names: Set[str] = set()
    for b in (args.baseline or []):
        for nm in load_name_list_from_file(b, args.baseline_column):
            baseline_names.add(nm)

    trusted_hosts = parse_hosts(args.trust_hosts)

    per_file_rows: List[Dict[str, object]] = []
    # If no explicit/global gold, build per-workbook gold from each workbook's survey sheet
    name_columns = [c.strip() for c in args.name_columns.split(";") if c.strip()]
    per_workbook_gold: Dict[str, List[str]] = {}
    use_per_workbook_gold = False
    if not gold_names:
        per_workbook_gold = build_per_workbook_gold(files, name_columns)
        use_per_workbook_gold = any(per_workbook_gold.values())
        if use_per_workbook_gold:
            print("Using per-workbook 'survey' sheet as gold for each workbook.")
    # Sort files so that within each workbook, methods are ordered: ours -> google -> datacite -> others
    def sort_key(p: str) -> Tuple[str, int, str]:
        wb = get_workbook_base(p) or p
        method = ""
        if "#" in p:
            method = p.rsplit("#", 1)[-1].lower()
        pr = 3
        if method in {"our", "ours"}:
            pr = 0
        elif method == "google":
            pr = 1
        elif method == "datacite":
            pr = 2
        return (wb, pr, method)

    files_sorted = sorted(files, key=sort_key)
    for p in files_sorted:
        # If using per-workbook gold, and this is a survey sheet, skip it from per-file output
        if use_per_workbook_gold and is_survey_input(p):
            continue
        wb_base = get_workbook_base(p)
        local_gold = gold_names
        if use_per_workbook_gold and wb_base and wb_base in per_workbook_gold:
            local_gold = per_workbook_gold.get(wb_base, [])
        s = evaluate_file(p, args, local_gold, baseline_names, trusted_hosts)
        per_file_rows.append(s)

    # Output per-file
    input_dir_name = os.path.basename(os.path.normpath(args.input_dir))
    per_file_path = os.path.join(args.output_dir, f"{input_dir_name}_per_file_v2.tsv")
    write_tsv(per_file_path, per_file_rows)

    # Aggregate micro
    agg = aggregate_micro(per_file_rows)
    agg_path = os.path.join(args.output_dir, f"{input_dir_name}_aggregate_v2.tsv")
    write_tsv(agg_path, [agg])

    print(f"Wrote: {per_file_path}")
    print(f"Wrote: {agg_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


